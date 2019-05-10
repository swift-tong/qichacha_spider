import json
import os
import re
import sys
import shutil
import time

from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
import requests


class MakeTransfer():
    def __init__(self):
        self.assets_list=[]
        self.column_list=[]
        self.time_list=[]
        self.union_list=[]
        self.excel_name=""
        self.url_dict={}
        self.change_dict={}
        self.name_new_dict={}
        self.merge_dict={}

    def set_property(self):
        self.dir_name=os.path.dirname(os.path.abspath(__file__))
        if os.path.exists('E:\接单\code_6'):
            self.excel_name = 'E:\接单\code_6\股东变更.xlsx'
            self.final_name=r'E:\接单\code_6\final\股东变更_{}.xlsx'.format(int(time.time()))
        else:
            self.excel_name='E:\接单\code_3\股东变更.xlsx'
            self.final_name = r'E:\接单\code_3\final\股东变更_{}.xlsx'.format(int(time.time()))
        self.asset_dir=os.path.join(self.dir_name,'assets')
        self.name_file=os.path.join(self.dir_name,'out','name_dict.json')
        self.name_new_file = os.path.join(self.dir_name, 'out', 'name_dict_new.json')
        print("asset_dir={}".format(self.asset_dir))

        with open(os.path.join(self.dir_name,'out','build_dict.json'),'r') as f:
            self.build_dict=json.load(f)
        print("build_dict={}".format(self.build_dict))

        tmp_dict={}
        with open(self.name_new_file,'r') as f:
            tmp_dict=json.load(f)
        print('tmp_dict={}'.format(tmp_dict))
        for key,value in tmp_dict.items():
            new_key=eval(key)
            self.name_new_dict[new_key]=value
        print(self.name_new_dict)

    def transfor_spider(self,key):
        url=self.url_dict[key]
        print('url={}'.format(url))

    def transfor_union(self,file,column_name):
        print("----------------Start transfor_union-------------")
        index_list=[]
        berfor_list=[]
        after_list=[]
        inter_list=[]#交集
        union_list=[]#
        union_dict={}
        lines=[]
        difference_a=[]#差集
        difference_b=[] # 差集
        with open(file,'r',encoding='utf-8') as f:
            lines=[va.replace('\n','') for va in f.readlines() if '带有*标记的为法定代表人	' not in va]
            for index,value in enumerate(lines):
                if '缴付期限' in value:
                    index_list.append(['缴付期限', index])
                elif '出资日期变更' in value:
                    index_list.append(['出资日期变更', index])
                elif '合伙人姓名或名称' in value:
                    index_list.append(['合伙人姓名或名称',index])
                elif '合伙人变更' in value:
                    index_list.append(['合伙人变更', index])
                elif '股东出资变更' in value:
                    index_list.append(['股东出资变更', index])
                elif '投资人' in value:
                    index_list.append(['投资人', index])
                elif '投资人(股权)变更' in value:
                    index_list.append(['投资人(股权)变更', index])
                elif '出资方式变更' in value:
                    index_list.append(['出资方式变更', index])
                elif '出资比例变更' in value:
                    index_list.append(['出资比例变更', index])
                elif '出资方式' in value:
                    index_list.append(['出资方式', index])
            print(index_list)
            for index ,value in enumerate(index_list):
                print(lines[value[1]])
                partten='.*\s+(\d+-\d+-\d+).*\w+.*'
                partten2 = '(\d+-\d+-\d+).*\w+.*'
                p1=re.search(partten,lines[value[1]])
                p2 = re.search(partten2, lines[value[1]])
                if p1:
                    time=p1.group(1)
                    print(time)
                elif p2:
                    time=p2.group(1)
                    print(time)
                else:
                    print("Have not get time!")
                    time=''
                start=value[1]+1
                if index < len(index_list) -1:
                    end=index_list[index+1][1]
                else:
                    end=len(lines)
                key=value[0]
                union_dict[(key,time)]=lines[start:end]
                print(start,end)
                print(union_dict)
                tmp_time='/'.join(time.split('-'))
                if column_name not in self.change_dict.keys():
                    self.change_dict[column_name]=[tmp_time]
                else:
                    value=self.change_dict[column_name]
                    if tmp_time not in value:
                        value.append(tmp_time)
                    value.sort()
                    self.change_dict[column_name]=value
            if column_name == "普华永道中天会计师事务所（特殊普通合伙）":
                print(column_name, union_dict)
            for key ,value in union_dict.items():
                if key[0] == '缴付期限' or key[0] == '出资日期变更':
                    tmp_list=[]
                    print(key,":",value)
                    for va in value:
                        print(va)
                        tmp_list.extend(va.split(';'))
                    union_dict[key]=tmp_list
                    print("After deal:")
                    print(key,union_dict[key])
            for key ,value in union_dict.items():
                if key[0] == '缴付期限' or key[0] == '出资日期变更':
                    for va in value:
                        p1 = re.search('(\D+)(\d+-\d+-\d+)\D+([\d | "."]+).*', va)
                        p2 = re.search('(\D+?)货币([\d | "."]+).*', va)
                        p3 = re.search('(\D+?)出资([\d | "."]+).*', va)
                        p4 = re.search('(\D+)(\d+-\d+-\d+)\D+.*', va)
                        if p1:
                            name = p1.group(1).strip()
                            time = p1.group(2)
                            amount = ''.join(p1.group(3).split(".")[:-1])
                            union_list.append((name, time, amount, '入股',column_name))
                        elif p2:
                            name = p2.group(1).strip()
                            time = key[1]
                            amount = p2.group(2)
                            union_list.append((name, time, amount, '入股',column_name))
                        elif p3:
                            name = p3.group(1).strip()
                            time = key[1]
                            amount = p3.group(2)
                            union_list.append((name, time, amount, '入股',column_name))
                        elif p4:
                            name = p4.group(1).strip()
                            time = p4.group(2)
                            amount = ''
                            union_list.append((name, time, amount, '入股',column_name))
                else:
                    for va in value:
                        p4 = re.search('\s*(\D+),(\d+).*', va)
                        p3=re.search('\s*(\D+).*货币.*?(\d+).*', va)
                        p5 = re.search('\s*(\D+).*?(\d+).*', va)
                        p1 = re.search('\s*(\D+)出资.*?(\d+).*', va)
                        p2 = re.search('\s*(\D+).*出资.*?(\d+).*', va)
                        if p1:
                            if '\n' in va:
                                va=va.replace('\n','')
                            if '【新增】' in va:
                                name=p1.group(1).strip()
                                time=key[1]
                                amount=p1.group(2)
                                union_list.append((name,time,amount,'新增',column_name))
                            elif '【退出】' in va:
                                name=p1.group(1).strip()
                                time=key[1]
                                amount=p1.group(2)
                                if '/'.join(time.split('-')) == self.change_dict[column_name][0]:
                                    union_list.append((name, self.build_dict[column_name], amount, '入股', column_name))
                                    union_list.append((name, time, amount, '退出', column_name))
                                else:
                                    union_list.append((name,time,amount,'退出',column_name))
                            else:
                                name=p1.group(1).strip()
                                time=key[1]
                                amount=p1.group(2)
                                if '/'.join(time.split('-')) == self.change_dict[column_name][0]:
                                    union_list.append((name, self.build_dict[column_name], amount, '入股', column_name))
                                    union_list.append((name, time, amount, '持股', column_name))
                                else:
                                    union_list.append((name, time, amount, '持股',column_name))
                        elif p2:
                            if '\n' in va:
                                va=va.replace('\n','')
                            if '【新增】' in va:
                                name=p2.group(1).strip()
                                time=key[1]
                                amount=p2.group(2)
                                union_list.append((name,time,amount,'新增',column_name))
                            elif '【退出】' in va:
                                name=p2.group(1).strip()
                                time=key[1]
                                amount=p2.group(2)
                                if '/'.join(time.split('-')) == self.change_dict[column_name][0]:
                                    union_list.append((name, self.build_dict[column_name], amount, '入股', column_name))
                                    union_list.append((name, time, amount, '退出', column_name))
                                else:
                                    union_list.append((name,time,amount,'退出',column_name))
                            else:
                                name=p2.group(1).strip()
                                time=key[1]
                                amount=p2.group(2)
                                if '/'.join(time.split('-')) == self.change_dict[column_name][0]:
                                    union_list.append((name, self.build_dict[column_name], amount, '入股', column_name))
                                    union_list.append((name, time, amount, '持股', column_name))
                                else:
                                    union_list.append((name, time, amount, '持股',column_name))
                        elif p3:
                            if '\n' in va:
                                va=va.replace('\n','')
                            if '【新增】' in va:
                                name=p3.group(1).strip()
                                time=key[1]
                                amount=p3.group(2)
                                union_list.append((name,time,amount,'新增',column_name))
                            elif '【退出】' in va:
                                name=p3.group(1).strip()
                                time=key[1]
                                amount=p3.group(2)
                                if '/'.join(time.split('-')) == self.change_dict[column_name][0]:
                                    union_list.append((name, self.build_dict[column_name], amount, '入股', column_name))
                                    union_list.append((name, time, amount, '退出', column_name))
                                else:
                                    union_list.append((name,time,amount,'退出',column_name))
                            else:
                                name=p3.group(1).strip()
                                time=key[1]
                                amount=p3.group(2)
                                if '/'.join(time.split('-')) == self.change_dict[column_name][0]:
                                    union_list.append((name, self.build_dict[column_name], amount, '入股', column_name))
                                    union_list.append((name, time, amount, '持股', column_name))
                                else:
                                    union_list.append((name, time, amount, '持股',column_name))
                        elif p4:
                            if '\n' in va:
                                va=va.replace('\n','')
                            if '【新增】' in va:
                                name=p4.group(1).strip()
                                time=key[1]
                                amount=p4.group(2)
                                union_list.append((name,time,amount,'新增',column_name))
                            elif '【退出】' in va:
                                name=p4.group(1).strip()
                                time=key[1]
                                amount=p4.group(2)
                                if '/'.join(time.split('-')) == self.change_dict[column_name][0]:
                                    union_list.append((name, self.build_dict[column_name], amount, '入股', column_name))
                                    union_list.append((name, time, amount, '退出', column_name))
                                else:
                                    union_list.append((name,time,amount,'退出',column_name))
                            else:
                                name=p4.group(1).strip()
                                time=key[1]
                                amount=p4.group(2)
                                if '/'.join(time.split('-')) == self.change_dict[column_name][0]:
                                    union_list.append((name, self.build_dict[column_name], amount, '入股', column_name))
                                    union_list.append((name, time, amount, '持股', column_name))
                                else:
                                    union_list.append((name, time, amount, '持股',column_name))
                        elif p5:
                            if '\n' in va:
                                va=va.replace('\n','')
                            if '【新增】' in va:
                                name=p5.group(1).strip()
                                time=key[1]
                                amount=p5.group(2)
                                union_list.append((name,time,amount,'新增',column_name))
                            elif '【退出】' in va:
                                name=p5.group(1).strip()
                                time=key[1]
                                amount=p5.group(2)
                                if '/'.join(time.split('-')) == self.change_dict[column_name][0]:
                                    union_list.append((name, self.build_dict[column_name], amount, '入股', column_name))
                                    union_list.append((name, time, amount, '退出', column_name))
                                else:
                                    union_list.append((name,time,amount,'退出',column_name))
                            else:
                                name=p5.group(1).strip()
                                time=key[1]
                                amount=p5.group(2)
                                if '/'.join(time.split('-')) == self.change_dict[column_name][0]:
                                    union_list.append((name, self.build_dict[column_name], amount, '入股', column_name))
                                    union_list.append((name, time, amount, '持股', column_name))
                                else:
                                    union_list.append((name, time, amount, '持股',column_name))
                        else:
                            p = re.search('(\D+).*', va)
                            if p:
                                if '\n' in va:
                                    va=va.replace('\n','')
                                if ';' in va:
                                    va=va.replace(';','')
                                if ',' in va:
                                    va=va.replace(',','')
                                    
                                if '【新增】' in va:
                                    name=va.replace('【新增】','').strip()
                                    time=key[1]
                                    amount=''
                                    union_list.append((name,time,amount,'新增',column_name))
                                elif '【退出】' in va:
                                    name=va.replace('【退出】','').strip()
                                    time=key[1]
                                    amount=''
                                    if '/'.join(time.split('-')) == self.change_dict[column_name][0]:
                                        union_list.append((name, self.build_dict[column_name], amount, '入股', column_name))
                                        union_list.append((name, time, amount, '退出', column_name))
                                    else:
                                        union_list.append((name,time,amount,'退出',column_name))
                                else:
                                    name=va.strip()
                                    time=key[1]
                                    amount=''
                                    print(sys._getframe().f_lineno,name)
                                    print(sys._getframe().f_lineno, time)
                                    if '/'.join(time.split('-')) == self.change_dict[column_name][0]:
                                        union_list.append((name, self.build_dict[column_name], amount, '入股', column_name))
                                        union_list.append((name, time, amount, '持股', column_name))
                                    else:
                                        union_list.append((name, time, amount, '持股', column_name))
                            else:
                                print("pass")
        print(union_list)
        self.assets_list.extend(union_list)

    def start_transfer(self):
        dir_list=os.listdir(self.asset_dir)
        for fi in dir_list:
            file=os.path.join(self.asset_dir,fi,'union.txt')
            self.union_list.append(file)
            column_name=os.path.basename(os.path.dirname(file))
            column_name=column_name.replace('（','(')
            column_name = column_name.replace('）', ')')
            print(column_name)
            self.column_list.append(column_name)
            self.transfor_union(file,column_name)
        print(self.change_dict)
        print("union_list={}".format(self.union_list))
        print("column_list={}".format(self.column_list))
        # sys.exit()
        # for item in self.column_list:
        #     self.transfor_spider(item)

    def write_excel2(self):
        wb=Workbook()
        ws=wb.active
        head_list=['姓名']
        tmp_list=[]
        name_list=[]
        for item in self.assets_list:
            if item[1] not  in tmp_list:
                tmp_list.append(item[1])
            if item[0] not in name_list:
                name_list.append([item[0]])
        sorted(tmp_list)
        head_list.extend(tmp_list)
        print(name_list)
        print("Totle {} people!".format(len(name_list)))
        ws.append(head_list)
        print(head_list)
        tmp_list2=['']*(len(head_list)-1)
        print(tmp_list2)
        for item2 in name_list:
            item2.extend(tmp_list2)
        print(len(head_list))
        print(len(name_list[0]))

        for item in self.assets_list:
            for na in name_list:
                print(na)
                print(item)
                if item[0] in na[0]:
                    head_index=head_list.index(item[1])
                    print(head_index)
                    print(item[3]+item[2])
                    na[head_index]=item[3]+item[2]
        for item in name_list:
            ws.append(item)
        wb.save(self.excel_name)

    def write_excel(self):
        head_list = [ '所在事务所','成立时间', '姓名','出资日期/缴付期限', '出资额', '撤资日期', '变更日期']
        if not os.path.exists(self.excel_name):
            excel=Workbook()
            ws = excel.active
            ws.append(head_list)
        else:
            excel=load_workbook(self.excel_name)
            ws = excel.active
        name_dict={}
        time_dict={}
        time_dict2 = {}
        name_list=[]
        for item in self.assets_list:
            name=item[0]
            time='/'.join(item[1].split('-'))
            assets_name=item[4]
            name_tuple=(name,assets_name)
            if assets_name not in time_dict.keys():
                time_dict[assets_name]=[time]
            else:
                value=time_dict[assets_name]
                if time not in value:
                    value.append(time)
                value.sort()
                time_dict[assets_name] = value
            if name_tuple not  in name_list:
                name_list.append(name_tuple)
            if name_tuple not in name_dict.keys():
                name_dict[name_tuple]=[]
        print("print(len(name_list))={}".format(len(name_list)))

        for key in name_dict.keys():
            tmp_list = []
            for item in self.assets_list:
                name = item[0]
                time = item[1]
                mony = item[2]
                type = item[3]
                column_name = item[4]
                if (name,column_name)==key:
                    if [time,mony,type] not in tmp_list:
                        tmp_list.append([time,mony,type])
                    tmp_list.sort()
            name_dict[key]=tmp_list
        print("name_dict={}".format(name_dict))
        print("len(name_dict)={}".format(len(name_dict)))
        new_dict={}
        for key,value in name_dict.items():
            try:
                new_dict[str(key)] = value
            except TypeError as e:
                print(e)
                print(key,value,sep=':')
        print('new_dict={}'.format(new_dict))
        print("len(new_dict)={}".format(len(new_dict)))
        json_str=json.dumps(new_dict,ensure_ascii=False,indent=4)
        with open(self.name_file,'w',encoding='utf-8') as f:
            f.write(json_str)

        print('before merge:')
        print('len(name_dict)={}'.format(len(name_dict)))
        print('len(name_new_dict)={}'.format(len(self.name_new_dict)))

        for key,value in name_dict.items():
            if key not  in self.merge_dict.keys():
                self.merge_dict[key]=value
            else:
                print('key {} already in merge_dict,will pass.'.format(key))
        new_key_dict = {}
        for key,value in self.name_new_dict.items():
            if key not  in self.merge_dict.keys():
                self.merge_dict[key]=value
            else:
                print('key {} already in merge_dict,will pass.'.format(key))

            if key[1] not in time_dict.keys():
                tmp_list = []
                for va in value:
                    if va[0] not in tmp_list:
                        tmp_list.append(va[0])
                tmp_list.sort()
                print(tmp_list)
                if key[1] not in new_key_dict.keys():
                    new_key_dict[key[1]] =tmp_list
                else:
                    value=new_key_dict[key[1]]
                    for tp in tmp_list:
                        if tp not in value:
                            value.append(tp)
                    value.sort()
                    new_key_dict[key[1]] = value
        for key,value in new_key_dict.items():
            if key not in self.change_dict.keys():
                self.change_dict[key] = value

        print('new_key_dict'.format(new_key_dict))
        for key,value in new_key_dict.items():
            if key not in time_dict.keys():
                time_dict[key]=value
        print('after merge:')
        print("time_dict={}".format(time_dict))
        time_str=json.dumps(time_dict,ensure_ascii=False,indent=4)
        time_file=os.path.join(self.dir_name,'out','time_dict.json')
        with open(time_file,'w',encoding='utf-8') as f:
            f.write(time_str)
        print('len(merge_dict)={}'.format(len(self.merge_dict)))
        for key,value in self.merge_dict.items():
            time_list = []
            name=key[0]
            for vax in value:
                timex='/'.join(vax[0].split('-'))
                if timex not in time_list:
                    time_list.append(timex)
                time_list.sort()
            time_dict2[name] = time_list
            final_list=['','','','','','','']
            head_list = [ '所在事务所','成立时间', '姓名','出资日期/缴付期限', '出资额', '撤资日期', '变更日期']
            final_list[0] = key[1]
            final_list[1] = self.build_dict[key[1]]
            final_list[2] = key[0]
            final_list[3] = time_list[0]
            ch_time=time_dict[key[1]]
            print('ch_time={}'.format(ch_time))
            print("time_list[-1]={}".format(time_list[-1]))
            print("time_dict[key[1]][-1]={}".format(time_dict[key[1]][-1]))
            if key[1] == "上会会计师事务所（特殊普通合伙）":
                print(time_list[-1],'XXX',self.change_dict[key[1]][-1])
            if time_list[-1] < self.change_dict[key[1]][-1]:
                final_list[5] = time_list[-1]
            final_list[6]=';'.join(ch_time)
            flag=True
            for va in value:
                if flag:
                    if '入股' in va:
                        final_list[3]='/'.join(va[0].split('-'))
                        print("final_list[2]={}".format(final_list[2]))
                        if va[1]:
                            final_list[4]="货币"+va[1]+"万人民币"
                        else:
                            final_list[4] =''
                        flag=False
                if '退出' in va:
                    final_list[5]='/'.join(va[0].split('-'))
            if key[0] == "TangTooPoh(陈图宝)":
                print(key[0], final_list)
            print(final_list)
            ws.append(final_list)
        json_str2 = json.dumps(time_dict2, ensure_ascii=False, indent=4)
        time_file2 = os.path.join(self.dir_name,'out', 'time_dict2.json')
        with open(time_file2, 'w', encoding='utf-8') as f:
            f.write(json_str2)
        print("time_dict2={}".format(time_dict2))
        excel.save(self.excel_name)
        shutil.copy(self.excel_name,self.final_name)


if __name__=="__main__":
    mt=MakeTransfer()
    mt.set_property()
    #make_folder()
    mt.start_transfer()
    print("assets_list:{}".format(mt.assets_list))
    print("len(mt.assets_list):{}".format(len(mt.assets_list)))
    mt.write_excel()
    #mt.transfor_first()