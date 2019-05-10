#coding=utf-8
import json
import os
import random
import shutil
import sys
import telnetlib
import urllib
from lxml import etree

import requests
import chardet
import time
import re
import xlwt
import xlrd
import datetime

from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options


class Excel_write():
    def __init__(self,path):
        self.path=path
        self.ex=xlwt.Workbook()

        pass


    def add_sheet(self,name,data_list,col=-1):

        sheet_list=[]
        if type(name)==list:
            for i in name:
                sh=self.ex.add_sheet(i, cell_overwrite_ok=True)
                if col!=-1:
                    self.set_width(sh,len(data_list),col)
                self.write(sh,0,data_list)

                sheet_list.append(sh)

            return sheet_list
        else:
            sh = self.ex.add_sheet(name, cell_overwrite_ok=True)
            if col != -1:
                self.set_width(sh, len(data_list), col)
            self.writes(sh, 0, data_list)
            return  sh



    def write(self, sheet, hang, data, index):
        for i in zip(index, data):
            sheet.write(hang, i[0], i[1])


    def writes(self,sheet,hang,data):
        for i in range(0,len(data)):
            sheet.write(hang,i,data[i])


    def set_width(self,sh,long_num,width):
        for i in range(0,long_num):
            sh.col(i).width = 256 * width

    def save(self):
        self.ex.save(self.path)

class MyRequests():
    def __init__(self):
        self.root_dir=os.path.dirname(os.path.abspath(__file__))
        self.no_partnerslist=[]
        self.parse_error_list=[]
        self.have_write_file=os.path.join(self.root_dir, 'out', 'have_write.json')
        with open(self.have_write_file,'r') as f:
            self.have_write_his=list(set(json.load(f)))
        print("Have_write_his={}".format(self.have_write_his))
        self.ok_lan=len(self.have_write_his)
        print(self.ok_lan)
        self.error_list=[]
        self.page_len=0
        self.current_page=0
        self.have_flag=False
        self.page_type='normal'
        self.build_dict={}
        self.count=0
        self.table_type=""
        self.build_time=""
        self.name_dict={}
        self.res_str=""
        self.proxie_list=[]
        self.have_write_current=[]
        with open(os.path.join(self.root_dir, 'out', 'change_list.json'),'r') as f:
            self.change_list=list(set(json.load(f)))
        print("change_list={}".format(self.change_list))
        self.ok_lan=len(self.change_list)
        print(self.ok_lan)
        self.multi_page=[]
        self.company_name=''
        self.excel_name=os.path.join(self.root_dir,'out','股东变更.xlsx')
        self.backup_file = os.path.join(self.root_dir,'backup','股东变更_{}.xlsx'.format(int(time.time())))
        self.name_file=os.path.join(self.root_dir,'out','name_dict.json')
        self.proxie_url='http://api3.xiguadaili.com/ip/?tid=556262668823744&num=5000&category=2&protocol=https&sortby=time'
        with open(os.path.join(self.root_dir,'config','user_agent.json'),'r') as f:
            self.agent_list=json.load(f)
        print(sys._getframe().f_lineno,'agent_list={}'.format(self.agent_list))
        self.headers = {
            'Cookie': 'QCCSESSID=i9blejlebmltbamlmakhkcgtt1; zg_did=%7B%22did%22%3A%20%2216a9744dc7529e-018ba9285fdd3d-3e385e0c-1fa400-16a9744dc76511%22%7D; UM_distinctid=16a9744dd1b38d-05b754053ce789-3e385e0c-1fa400-16a9744dd1c3ed; _uab_collina=155731603430493590323018; acw_tc=6f48649915573158844101416e20524a4b8cb4a73be827555c917da8d5; zg_de1d1a35bfa24ce29bbf2c7eb17e6c4f=%7B%22sid%22%3A%201557492685678%2C%22updated%22%3A%201557492686718%2C%22info%22%3A%201557316033670%2C%22superProperty%22%3A%20%22%7B%7D%22%2C%22platform%22%3A%20%22%7B%7D%22%2C%22utm%22%3A%20%22%7B%7D%22%2C%22referrerDomain%22%3A%20%22%22%2C%22cuid%22%3A%20%224f7112a3d7c45eb79a3965e430e5943b%22%7D; CNZZDATA1254842228=1330679215-1557312612-https%253A%252F%252Fwww.baidu.com%252F%7C1557490909; Hm_lvt_3456bee468c83cc63fb5147f119f1075=1557317072,1557403371,1557487605,1557492688; Hm_lpvt_3456bee468c83cc63fb5147f119f1075=1557492688',
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.103 Safari/537.36',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language':'zh-CN,zh;q=0.9',
            'Accept':'text/html, */*; q=0.01',
            'Referer': 'https://www.qichacha.com/',
            'X-Requested-With': 'XMLHttpRequest',
            'Connection': 'keep-alive',
            'Content-Type': 'application/x-www-form-urlencoded'
            }

        with open(os.path.join(self.root_dir,'config','current_url.json'),'r',encoding='gbk') as f:
            self.company_url=json.load(f)
        print(sys._getframe().f_lineno,'company_url={}'.format(self.company_url))

    def check_ip(self,ipx):
        print(sys._getframe().f_lineno,"Will check {}".format(ipx))
        header = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36",
        }
        try:
            res= requests.get("http://www.baidu.com/", headers=header, proxies={"https": "https://{}".format(ipx)}, timeout=1)
            if res.status_code == 200:
                print(sys._getframe().f_lineno,"该代理IP可用：", ipx)
                return True
            else:
                print(sys._getframe().f_lineno,"该代理IP不可用：", ipx)
                return False
        except Exception as  e:
            print(sys._getframe().f_lineno,e)
            print(sys._getframe().f_lineno,"IP can not use!")
            return False

    def verification(self,respnse):
        html=etree.HTML(respnse.text)
        print(etree.fromstring(html))
        span=html.xpath('//*[@id="nc_1_n1z"]')
        print(span)

    # proxies ={'https':"127.0.0.1:8888} or proxies ={'http':"127.0.0.1:8888}
    def get(self,url,headers={},timeout=60,proxies=None,params={}):
        print(sys._getframe().f_lineno,"proxies={}".format(proxies))
        print(sys._getframe().f_lineno,"Will get : {}".format(url))
        headers['User-Agent']=random.choice(self.agent_list)
        print(sys._getframe().f_lineno, "headers['Cookie']={}".format(self.headers['Cookie']))
        print(sys._getframe().f_lineno,'params={}'.format(params))
        try:
            sleep_time=random.randrange(2,10)
            print("Will sleep {} second!".format(sleep_time))
            time.sleep(sleep_time)
            r=requests.get(url=url,headers=headers,timeout=timeout,params=params)
            print(sys._getframe().f_lineno,chardet.detect(r.content))
            print(sys._getframe().f_lineno,type(r.content))
            print(sys._getframe().f_lineno,'response  responce.headers={}'.format(r.headers))
            print(sys._getframe().f_lineno,"response  status_code:",type(r.status_code), r.status_code)
            print(sys._getframe().f_lineno,"response  headers:",type(r.headers), r.headers)
            print(sys._getframe().f_lineno,"response  cookies:",type(r.cookies), r.cookies)
            print(sys._getframe().f_lineno,"response  url:",type(r.url), r.url)
            print(sys._getframe().f_lineno,"response  history:",type(r.history), r.history)
            encoding=chardet.detect(r.content)['encoding']
            r.encoding=encoding
            if r.status_code==200:
                if "text/html" in r.headers['Content-Type']:
                    print(sys._getframe().f_lineno,'"text/html" in r.headers,return r.text')
                    return r.text
                elif "application/json;" in r.headers['Content-Type']:
                    print(sys._getframe().f_lineno,'"application/json;" in r.headers,return r.text')
                    return r.json()
            else:
                print(sys._getframe().f_lineno,'get reponse false!')
                print(sys._getframe().f_lineno,r.status_code)
                print(sys._getframe().f_lineno,r.reason)
                return None

        except Exception as e:
            print(sys._getframe().f_lineno,e)
            self.save_data()
            sys.exit()
            return None

    def get_data(self,url):
        print("enter get_data")
        try:
            uni = url.split('_')[-1].split('.')[0]
            print(sys._getframe().f_lineno,'unique={}'.format(uni))
            self.assert_name = etree.HTML(self.res_str).xpath('//*[@id="company-top"]/div[2]/div[2]/div[1]/h1/text()')[0]
            print(sys._getframe().f_lineno,'assert_name={}'.format(self.assert_name))
            companyname=urllib.parse.quote(self.assert_name)
            print(sys._getframe().f_lineno,'companyname={}'.format(companyname))
            ul_tr=etree.HTML(self.res_str).xpath('//ul[@class="pagination"]/li/a/text()')
            print(sys._getframe().f_lineno,'ul_tr={}'.format(ul_tr))
            if not ul_tr:
                print(sys._getframe().f_lineno,"Only one page,Will return None!")
                return None
            number_list=[]
            for item in ul_tr:
                p=re.search('.*(\d+).*',item)
                if p:
                    number_list.append(p.group(1))
                else:
                    print(sys._getframe().f_lineno,'item is not number')
            print(sys._getframe().f_lineno,number_list)
            print(sys._getframe().f_lineno,'len(ul_tr)={}'.format(len(ul_tr)))
            data = {
                'unique': uni,
                'companyname': companyname,
                'p': number_list,
                'tab': 'base',
                'box': 'partners'
            }
            print(sys._getframe().f_lineno,'data={}'.format(data))
            return data
        except Exception as  e:
            print(sys._getframe().f_lineno,e)
            self.save_data()
            sys.exit()

    def process_parse(self):
        j=""
        print(sys._getframe().f_lineno,'In process_parse')
        print(self.res_str)
        try:
            # print(self.res_str)
            if self.page_type == "normal":
                tr1 = etree.HTML(self.res_str).xpath('//section[@id="partnerslist"]/table[@class="ntable ntable-odd ntable-stext"]/tr')
            elif self.page_type == "first":
                tr1 = etree.HTML(self.res_str).xpath('//section[@id="partnerslist"]/table[@class="ntable ntable-odd "]/tr')
            elif self.page_type == "multi":
                tr1 = etree.HTML(self.res_str).xpath('//section[@id="partnerslist"]/table[@class="ntable ntable-odd "]/tbody/tr')[1:]
            else:
                print("xpath error,please recheck!")
                sys.exit()
            tr2=[]
            for t in tr1:
                str_tr=etree.tostring(t,pretty_print=True)
                print(str_tr)
                if 'th class="tx"' not in str_tr.decode('utf-8'):
                    tr2.append(t)
            if tr1==tr2:
                print('tr1==tr2')
            print("len(tr2)={}".format(len(tr2)))
            # with open("test.html",'w',encoding='utf-8') as  f:
            #     f.write(self.res_str)
            for i in range(0,len(tr2)):
                j=tr2[i]
                mon = []
                name = j.xpath('./td[2]/table//tr/td[2]/a/h3/text()')[0]
                print(sys._getframe().f_lineno,'name={}'.format(name))
                money = j.xpath('./td[4]/text()')[0].replace('\n', '').replace('\t', '').replace(' ', '')+"万元人民币"
                print(sys._getframe().f_lineno,'money={}'.format(money))
                time = j.xpath('./td[5]/text()')[0].replace('\n', '').replace('\t', '').replace(' ', '')
                print(sys._getframe().f_lineno,'time={}'.format(time))
                p = re.search('(\d+-\d+-\d+).*', time)
                if p:
                    time = '/'.join(p.group(1).split('-'))
                else:
                    print(sys._getframe().f_lineno,'Get time ={},will set the time as assets build time.'.format(time))
                    time = '/'.join(self.build_time.split('-'))
                    print(sys._getframe().f_lineno,'xxx time={}'.format(time))
                print(sys._getframe().f_lineno,'new time={}'.format(time))
                mon=[time, money,'入股']
                key=str((name,self.company_name))
                print(sys._getframe().f_lineno,key)
                if key not in self.name_dict.keys():
                    self.name_dict[key]=[mon]
                else:
                    value=self.name_dict[key]
                    print(sys._getframe().f_lineno,'value={}'.format(value))
                    print(sys._getframe().f_lineno,'mon={}'.format(mon))
                    if mon not in value:
                        value.extend(mon)
                        self.name_dict[key]=value
                info_list=[self.company_name,self.build_time,name,time,money,'','']
                print('info_list={}'.format(info_list))
                if self.company_name not in self.have_write_current:
                    self.have_write_current.append(self.company_name)
                self.write_excel(info_list)
        except Exception as  e:
            print(sys._getframe().f_lineno,"In process_parse,parse {} error!".format(self.company_name))
            print(e)
            print(tr2)
            print(etree.tostring(j,pretty_print=True))
            self.save_data()
            if self.company_name not in self.parse_error_list:
                self.parse_error_list.append(self.company_name)
            print(sys._getframe().f_lineno,e)

    def write_excel(self,info_list):
        head_list = [ '所在事务所','成立时间', '姓名','出资日期/缴付期限', '出资额', '撤资日期', '变更日期']
        if not os.path.exists(self.excel_name):
            excel=Workbook()
            ws = excel.active
            ws.append(head_list)
            ws.append(info_list)
        else:
            print(sys._getframe().f_lineno,"Will write {} to excel!".format(info_list))
            excel=load_workbook(self.excel_name)
            ws = excel.active
            ws.append(info_list)
        excel.save(self.excel_name)
        shutil.copy(self.excel_name, self.backup_file)

    def save_data(self):
        print(sys._getframe().f_lineno,'In save_data,will write data to file! ')

        build_dict=json.dumps(self.build_dict,ensure_ascii=False,indent=4)
        fi_hui=os.path.join(self.root_dir,'out','build_dict.json')
        with open(fi_hui,'w') as  f:
            f.write(build_dict)
        return True

        str_dict=json.dumps(self.name_dict,ensure_ascii=False,indent=4)
        with open(self.name_file,'w') as  f:
            f.write(str_dict)

        str_list=json.dumps(self.no_partnerslist,ensure_ascii=False,indent=4)
        fi=os.path.join(self.root_dir,'error','no_partnerslist.json')
        with open(fi,'w') as  f:
            f.write(str_list)

        change_list=json.dumps(self.change_list,ensure_ascii=False,indent=4)
        fi=os.path.join(self.root_dir,'out','change_list.json')
        with open(fi,'w') as  f:
            f.write(change_list)

        multi_page=json.dumps(self.multi_page,ensure_ascii=False,indent=4)
        fi=os.path.join(self.root_dir,'out','multi_page.json')
        with open(fi,'w') as  f:
            f.write(multi_page)

        str_list2=json.dumps(self.parse_error_list,ensure_ascii=False,indent=4)
        fi2=os.path.join(self.root_dir,'error','parse_error.json')
        with open(fi2,'w') as f:
            f.write(str_list2)

        str_list3=json.dumps(self.error_list,ensure_ascii=False,indent=4)
        fi3=os.path.join(self.root_dir,'error','error_list.json')
        with open(fi3,'w') as f:
            f.write(str_list3)

        self.have_write_his.extend(self.have_write_current)
        self.have_write_his=list(set(self.have_write_his))
        str_list4=json.dumps(self.have_write_his,ensure_ascii=False,indent=4)
        print("have_write_his={}".format(self.have_write_his))
        print("len(have_write_his)={}".format(len(self.have_write_his)))
        print("have_write_current={}".format(self.have_write_current))
        with open(self.have_write_file,'w') as f:
            f.write(str_list4)
        if self.have_flag:
            flag=os.path.join(self.root_dir,"run")
            with open(flag, "w") as f:
                f.write("xx")

    def check_partnerslist(self,key,url):
        build_time =etree.HTML(self.res_str).xpath('//section[@id="Cominfo"]/table[2]/tr[2]/td[4]/text()')[0].replace('\n', '').replace('\t', '').replace(' ', '')
        self.build_time='/'.join(build_time.split('-'))
        print(sys._getframe().f_lineno,'build_time={}'.format(self.build_time))
        company_name=etree.HTML(self.res_str).xpath('//*[@id="company-top"]/div[2]/div[2]/div[1]/h1/text()')[0]
        self.company_name=company_name
        print(sys._getframe().f_lineno,'company_name={}'.format(company_name))
        partnerslist = etree.HTML(self.res_str).xpath('//*[@id="partnerslist"]')
        print(sys._getframe().f_lineno,partnerslist)
        if len(partnerslist) > 0:
            print(sys._getframe().f_lineno,"partnerslist exists! Will check table type!")
            ret2=partnerslist[0].xpath('./table[@class="ntable ntable-odd "]/tr[2]/td')
            lenx=len(ret2)
            print(sys._getframe().f_lineno,"len(ret2)={}".format(lenx))
            if lenx > 5:
                self.table_type="type7"
            else:
                self.table_type="type5"
            print(sys._getframe().f_lineno,"table_type={}".format(self.table_type))
            return True
        else:
            if (key, url) not in self.no_partnerslist:
                self.no_partnerslist.append((key, url))
            # out_html=os.path.join(self.root_dir,'out',key+'.html')
            # with open(out_html,'w',encoding='utf-8') as f:
            #     f.write(etree.tostring(etree.fromstring(res)))
            return False

    def check_change_list(self):
        print('Enter check_change_list')
        changelist=''
        print(sys._getframe().f_lineno, "Will create build_dict")
        print(sys._getframe().f_lineno, "build time={}".format(self.build_time))
        if self.company_name not in self.build_dict.keys():
            self.build_dict[self.company_name] = self.build_time
        try:
            changelist=etree.HTML(self.res_str).xpath('//*[@id="Changelist"]')[0]
            if len(changelist) > 0:
                print(sys._getframe().f_lineno,"{} have change list!".format(self.company_name))
                if self.company_name not in self.change_list:
                    self.change_list.append(self.company_name)
                # print(sys._getframe().f_lineno, "Will create build_dict")
                # print(sys._getframe().f_lineno, "build time={}".format(self.build_time))
                # if self.company_name not in self.build_dict.keys():
                #     self.build_dict[self.company_name]=self.build_time
        except Exception as e:
            print(sys._getframe().f_lineno,"Get changelist Failed!")
            print(sys._getframe().f_lineno,e)


    def check_table(self,key,url):
        print(sys._getframe().f_lineno,'In check_table')
        try:
            self.res_str = self.get(url=url, headers=self.headers)
            ret=self.check_partnerslist(key,url)
            # if ret:
            self.check_change_list()
            return ret
        except Exception as  e:
            print(sys._getframe().f_lineno,"exception happened in check_table,please recheck!")
            print(sys._getframe().f_lineno,e)
            if self.company_name not in self.error_list.append:
                self.error_list.append(self.company_name)
            print(sys._getframe().f_lineno,e)
            self.save_data()
            sys.exit()
            self.count=self.count+1
            if self.count >10:
                print("Some wrong happen,Will exit")
                self.save_data()
                return "exit"

    def myselenium2(self,url):
        url=url
        chrome_option=Options()
        chrome_option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        executable_path="C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
        os.environ["webdriver.chrome.driver"]=r'C:\Users\Administrator\AppData\Local\Google\Chrome\Application'
        driver=webdriver.Chrome(chrome_options=chrome_option)
        print(driver.title)
        # self.driver_celcle(driver)
        driver.get(url)
        driver.implicitly_wait(20)
        size = driver.get_window_size()
        print(size)
        page=[]
        pa = driver.find_elements_by_xpath('//*[@id="ajaxpage"]')
        for p in pa:
            if p.text.isdigit():
                page.append(p)
        self.page_len=len(page)
        print("page length {}".format(self.page_len))

        pa=page[self.current_page]
        tx=pa.text
        pa.click()
        time.sleep(2)
        new_pa = driver.find_elements_by_xpath('//*[@id="partnerslist"]/div[2]/nav/ul//li/a')
        for new_p in new_pa:
            if new_p.text == tx:
                print(new_p.id)
                print(new_p.tag_name)
                print(new_p.text)
                self.res_str=driver.page_source
                self.page_type='multi'
                self.process_parse()
        self.current_page=self.current_page+1
        if self.current_page > self.page_len-1:
            return True
        else:
            self.myselenium2(url)

    def process_run(self):
        data=""
        # self.proxie_list=requests.get(self.proxie_url).text.split('\r\n')
        # print(sys._getframe().f_lineno,"proxie_list={}".format(self.proxie_list))
        # print(sys._getframe().f_lineno,type(self.proxie_list))
        for key,value in self.company_url.items():
        # for key, value in [('中瑞岳华会计师事务所(特殊普通合伙)','https://www.qichacha.com/firm_84bca9d6d559d89882a0721b5f91e6d9.html')]:
            print("----------------------------------------------------")
            # if key not in self.have_write_his:
            if True:
                self.current_page = 0
                print('Will get {}'.format(key))
                time.sleep(random.randrange(1, 2))
                self.have_flag=True
                ret= self.check_table(key,value)
                continue
                print(sys._getframe().f_lineno,'check_table ret ={}'.format(ret))
                if not ret:
                    print(sys._getframe().f_lineno,"{} have no 股东信息，will pass it！".format(key))
                    continue
                if ret == "exit":
                    return ret
                else:
                    data=self.get_data(value)
                if data:
                    self.multi_page.append(key)
                    self.page_type='first'
                    self.process_parse()
                    self.myselenium2(value)
                else:
                    print(sys._getframe().f_lineno,'Will deal one page company!')
                    self.page_type='normal'
                    self.process_parse()
            else:
                print("Pass {}".format(key))
        self.save_data()


if __name__ == '__main__':
    start = time.time()
    mr=MyRequests()
    ret=mr.process_run()
    end=time.time()
    use_time=(end-start)/60
    print(sys._getframe().f_lineno,"Totle use {} minutes".format(use_time))



