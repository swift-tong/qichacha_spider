import json
import os
import re
import shutil
import sys
import shutil
import time

from openpyxl import Workbook, load_workbook


class MakeTransfer():
    def __init__(self):
        self.assets_list=[]
        self.type1_list=[]
        self.type2_list=[]
        self.type3_list = []
        self.type4_list = []
        self.time_list=[]
        self.excel_name=""

    def set_excel_name(self):
        dir_name=os.path.dirname(os.path.abspath(__file__))
        if os.path.exists('E:\workspace\接单\code_3'):
            self.excel_name = 'E:\workspace\接单\code_3\股东变更.xlsx'
            self.final_name=r'E:\workspace\接单\code_3\final\股东变更_{}.xlsx'.format(int(time.time()))
        else:
            self.excel_name='E:\接单\code_3\股东变更.xlsx'
            self.final_name = r'E:\接单\code_3\final\股东变更_{}.xlsx'.format(int(time.time()))
        self.asset_name=os.path.basename(dir_name)
        print(self.excel_name)

    def get_type_list(self):
        filist=os.listdir('.')
        for fi in filist:
            if fi.endswith('缴付期限.txt'):
                pass
                self.type1_list.append(fi)
            elif fi.endswith('合伙人姓名或名称.txt') or fi.endswith('股东出资变更.txt')or fi.endswith('投资人.txt'):
                time=fi.split('_')[0]
                self.type2_list.append(fi)
            elif fi.endswith('出资方式.txt'):
                time=fi.split('_')[0]
                self.type3_list.append(fi)
            elif fi.endswith('union.txt'):
                time=fi.split('_')[0]
                self.type4_list.append(fi)
            else:
                print('pass file {}'.format(fi))
        print("type1_list:")
        print(self.type1_list)
        print("type2_list:")
        print(self.type2_list)
        print("type3_list:")
        print(self.type3_list)
        print("change_list:")

    def transfor_first(self,file1):
        print("----------------Start transfor_first-------------")
        berfor_list=[]
        after_list=[]
        inter_list=[]#交集
        union_list=[]#并集
        difference_a=[]#差集
        difference_b=[] # 差集
        with open(file1,'r',encoding='utf-8') as f:
            lines=f.readlines()
            n=len(lines)
            for item in lines[0].split(";"):
                p=re.search('(\D+)(\d+-\d+-\d+)\D+([\d | "."]+).*',item)
                if p:
                    name=p.group(1)
                    time=p.group(2)
                    amount=''.join(p.group(3).split(".")[:-1])
                    berfor_list.append((name,time,amount,'入股'))
                else:
                    print("Not find!")
            print(berfor_list)
            print("berfor length:",len(berfor_list))
            for item in lines[1].split(";"):
                p=re.search('(\D+)(\d+-\d+-\d+)\D+([\d | "."]+).*',item)
                if p:
                    name=p.group(1)
                    time=p.group(2)
                    amount=''.join(p.group(3).split(".")[:-1])
                    after_list.append((name,time,amount,'入股'))
                else:
                    print("Not find!")
            print(after_list)
            print("after_list length:", len(after_list))
            difference_a=set(berfor_list).difference(set(after_list))
            difference_b = set(after_list).difference(set(berfor_list))
            inter_list=set(berfor_list).intersection(set(after_list))
            union_list=set(berfor_list).union(set(after_list))
            # print(difference_a)
            # print(difference_b)
            print(inter_list)
            print("inter_list length:", len(inter_list))
            print(union_list)
            print("union_list length:", len(union_list))
            self.assets_list.extend(union_list)

    def transfor_second(self,file2,time):
        print("----------------Start transfor_second-------------")
        berfor_list=[]
        after_list=[]
        inter_list=[]#交集
        union_list=[]#并集
        difference_a=[]#差集
        difference_b=[] # 差集
        with open(file2,'r',encoding='utf-8') as f:
            lines=f.readlines()
            for na in lines:
                p = re.search('(\D+),(\d+).*', na)
                if p:
                    if '\n' in na:
                        na=na.replace('\n','')
                    if '【新增】' in na:
                        name=p.group(1)
                        time=time
                        amount=p.group(2)
                        union_list.append((name,time,amount,'新增'))
                    elif '【退出】' in na:
                        name=p.group(1)
                        time=time
                        amount=p.group(2)
                        union_list.append((name,time,amount,'退出'))
                    else:
                        name=p.group(1)
                        time=time
                        amount=p.group(2)
                        union_list.append((name, time, amount, '持股'))
                else:
                    if '\n' in na:
                        na=na.replace('\n','')
                    if '【新增】' in na:
                        name=na.replace('【新增】','')
                        time=time
                        amount=''
                        union_list.append((name,time,amount,'新增'))
                    elif '【退出】' in na:
                        name=na.replace('【退出】','')
                        time=time
                        amount=''
                        union_list.append((name,time,amount,'退出'))
                    else:
                        name=na
                        time=time
                        amount=''
                        union_list.append((name, time, amount, '持股'))
        print(union_list)
        self.assets_list.extend(union_list)

    def transfor_third(self,file3,time):
        print("----------------Start transfor_third-------------")
        berfor_list=[]
        after_list=[]
        inter_list=[]#交集
        union_list=[]#并集
        difference_a=[]#差集
        difference_b=[] # 差集
        line_list=[]
        new_lines=[]
        new_lines2=[]
        new_lines3 = []
        with open(file3,'r',encoding='utf-8') as f:
            lines=f.readlines()

        for index,item in enumerate(lines):
            # lines[index]=re.sub(';\s*[新增]','[新增]',item)
            # lines[index] = re.sub(';\s*[退出]', '[退出]', item)
            tmp_list=lines[index].split(';')
            new_lines.extend(tmp_list)
        for item in new_lines:
            if '\n' in item:
                new_lines2.append(item.replace('\n',''))
            else:
                new_lines2.append(item)
        for item in new_lines2:
            if item != ' ':
                new_lines3.append(item)
        print(new_lines3)
        print(len(new_lines3))
        for na in new_lines3:
            p=re.search('(\D+)(\d+).*',na)
            p2=re.search('(\D+).*[退出].*',na)
            p3 = re.search('(\D+).*[新增].*', na)
            if p or p2 or p3:
                if '[新增]' in na:
                    name = p3.group(1).replace('货币','') if '货币' in p3.group(1) else p3.group(1)
                    time = time
                    amount = 50
                    line_list.append((name, time, amount, '新增'))
                elif '[退出]' in na:
                    name =  p2.group(1).replace('货币','') if '货币' in p2.group(1) else p2.group(1)
                    time = time
                    amount =  50
                    line_list.append((name, time, amount, '退出'))
                else:
                    name = p.group(1).replace('货币','') if '货币' in p.group(1) else p.group(1)
                    time = time
                    amount = p.group(2)
                    line_list.append((name, time, amount, '持股'))
            else:
                print('In transfor_third,can not deal {}'.format(na))
        print("Transfor_third line list {}".format(line_list))
        self.assets_list.extend(line_list)

    def transfor_union(self,file4):
        print("----------------Start transfor_union-------------")
        index_list=[]
        berfor_list=[]
        after_list=[]
        inter_list=[]#交集
        union_list=[]#
        union_dict={}
        lines=[]
        difference_a=[]#差集
        difference_b=[] # 差集
        with open(file4,'r',encoding='utf-8') as f:
            lines=[va.replace('\n','') for va in f.readlines() if '带有*标记的为法定代表人	' not in va]
            for index,value in enumerate(lines):
                if '缴付期限' in value:
                    index_list.append(['缴付期限', index])
                elif '出资日期变更' in value:
                    index_list.append(['出资日期变更', index])
                elif '合伙人姓名或名称' in value:
                    index_list.append(['合伙人姓名或名称',index])
                elif '合伙人变更' in value:
                    index_list.append(['合伙人变更', index])
                elif '股东出资变更' in value:
                    index_list.append(['股东出资变更', index])
                elif '投资人' in value:
                    index_list.append(['投资人', index])
                elif '投资人(股权)变更' in value:
                    index_list.append(['投资人(股权)变更', index])
                elif '出资方式变更' in value:
                    index_list.append(['出资方式变更', index])
                elif '出资比例变更' in value:
                    index_list.append(['出资比例变更', index])
            print(index_list)
            for index ,value in enumerate(index_list):
                print(lines[value[1]])
                partten='\d*\s*(\d+-\d+-\d+).*\w+.*'
                p=re.search(partten,lines[value[1]])
                if p:
                    time=p.group(1)
                    print(time)
                else:
                    print("Have not get time!")
                    time=''
                start=value[1]+1
                if index < len(index_list) -1:
                    end=index_list[index+1][1]
                else:
                    end=len(lines)
                key=value[0]
                union_dict[(key,time)]=lines[start:end]
                print(start,end)
                print(union_dict)
            for key ,value in union_dict.items():
                if key[0] == '缴付期限' or key[0] == '出资日期变更':
                    tmp_list=[]
                    print(key,":",value)
                    for va in value:
                        print(va)
                        tmp_list.extend(va.split(';'))
                    union_dict[key]=tmp_list
                    print("After deal:")
                    print(key,union_dict[key])
            for key ,value in union_dict.items():
                if key[0] == '缴付期限' or key[0] == '出资日期变更':
                    for va in value:
                        p1 = re.search('(\D+)(\d+-\d+-\d+)\D+([\d | "."]+).*', va)
                        p2 = re.search('(\D+?)货币([\d | "."]+).*', va)
                        p3 = re.search('(\D+?)出资([\d | "."]+).*', va)
                        p4 = re.search('(\D+)(\d+-\d+-\d+)\D+.*', va)
                        if p1:
                            name = p1.group(1)
                            time = p1.group(2)
                            amount = ''.join(p1.group(3).split(".")[:-1])
                            union_list.append((name, time, amount, '入股'))
                        elif p2:
                            name = p2.group(1)
                            time = key[1]
                            amount = p2.group(2)
                            union_list.append((name, time, amount, '入股'))
                        elif p3:
                            name = p3.group(1)
                            time = key[1]
                            amount = p3.group(2)
                            union_list.append((name, time, amount, '入股'))
                        elif p4:
                            name = p4.group(1)
                            time = p4.group(2)
                            amount = ''
                            union_list.append((name, time, amount, '入股'))
                else:
                    for va in value:
                        p1 = re.search('(\D+),(\d+).*', va)
                        p2=re.search('(\D+).*?货币.*?(\d+).*', va)
                        p3 = re.search('(\D+).*?(\d+).*', va)
                        p4 = re.search('(\D+).*?出资.*?(\d+).*', va)
                        if p1:
                            if '\n' in va:
                                va=va.replace('\n','')
                            if '【新增】' in va:
                                name=p1.group(1)
                                time=key[1]
                                amount=p1.group(2)
                                union_list.append((name,time,amount,'新增'))
                            elif '【退出】' in va:
                                name=p1.group(1)
                                time=key[1]
                                amount=p1.group(2)
                                union_list.append((name,time,amount,'退出'))
                            else:
                                name=p1.group(1)
                                time=key[1]
                                amount=p1.group(2)
                                union_list.append((name, time, amount, '持股'))
                        elif p2:
                            if '\n' in va:
                                va=va.replace('\n','')
                            if '【新增】' in va:
                                name=p2.group(1)
                                time=key[1]
                                amount=p2.group(2)
                                union_list.append((name,time,amount,'新增'))
                            elif '【退出】' in va:
                                name=p2.group(1)
                                time=key[1]
                                amount=p2.group(2)
                                union_list.append((name,time,amount,'退出'))
                            else:
                                name=p2.group(1)
                                time=key[1]
                                amount=p2.group(2)
                                union_list.append((name, time, amount, '持股'))
                        elif p3:
                            if '\n' in va:
                                va=va.replace('\n','')
                            if '【新增】' in va:
                                name=p3.group(1)
                                time=key[1]
                                amount=p3.group(2)
                                union_list.append((name,time,amount,'新增'))
                            elif '【退出】' in va:
                                name=p3.group(1)
                                time=key[1]
                                amount=p3.group(2)
                                union_list.append((name,time,amount,'退出'))
                            else:
                                name=p3.group(1)
                                time=key[1]
                                amount=p3.group(2)
                                union_list.append((name, time, amount, '持股'))
                        elif p4:
                            if '\n' in va:
                                va=va.replace('\n','')
                            if '【新增】' in va:
                                name=p4.group(1)
                                time=key[1]
                                amount=p4.group(2)
                                union_list.append((name,time,amount,'新增'))
                            elif '【退出】' in va:
                                name=p4.group(1)
                                time=key[1]
                                amount=p4.group(2)
                                union_list.append((name,time,amount,'退出'))
                            else:
                                name=p4.group(1)
                                time=key[1]
                                amount=p4.group(2)
                                union_list.append((name, time, amount, '持股'))
                        else:
                            p = re.search('(\D+).*', va)
                            if p:
                                if '\n' in va:
                                    va=va.replace('\n','')
                                if '【新增】' in va:
                                    name=va.replace('【新增】','')
                                    time=key[1]
                                    amount=''
                                    union_list.append((name,time,amount,'新增'))
                                elif '【退出】' in va:
                                    name=va.replace('【退出】','')
                                    time=key[1]
                                    amount=''
                                    union_list.append((name,time,amount,'退出'))
                                else:
                                    name=va
                                    time=key[1]
                                    amount=''
                                    union_list.append((name, time, amount, '持股'))
                            else:
                                print("pass")
        for item in union_list:
            if item[1] not in self.time_list:
                self.time_list.append('/'.join(item[1].split('-')))
        print(union_list)
        self.assets_list.extend(union_list)

    def start_transfer(self):
        self.get_type_list()
        if len(self.type1_list) >0:
            for fi1 in self.type1_list:
                self.transfor_first(fi1)

        if len(self.type2_list) > 0:
            for fi2 in self.type2_list:
                time=fi2.split('_')[0]
                self.time_list.append(time)
                # new_time=time[:4]+"-"+time[4:6]+"-"+time[6:]
                self.transfor_second(fi2,time)

        if len(self.type3_list) > 0:
            for fi3 in self.type3_list:
                time=fi3.split('_')[0]
                self.time_list.append(time)
                self.transfor_third(fi3,time)

        if len(self.type4_list) > 0:
            for fi4 in self.type4_list:
                self.transfor_union(fi4)

    def write_excel2(self):
        wb=Workbook()
        ws=wb.active
        head_list=['姓名']
        tmp_list=[]
        name_list=[]
        for item in self.assets_list:
            if item[1] not  in tmp_list:
                tmp_list.append(item[1])
            if item[0] not in name_list:
                name_list.append([item[0]])
        sorted(tmp_list)
        head_list.extend(tmp_list)
        print(name_list)
        print("Totle {} people!".format(len(name_list)))
        ws.append(head_list)
        print(head_list)
        tmp_list2=['']*(len(head_list)-1)
        print(tmp_list2)
        for item2 in name_list:
            item2.extend(tmp_list2)
        print(len(head_list))
        print(len(name_list[0]))

        for item in self.assets_list:
            for na in name_list:
                print(na)
                print(item)
                if item[0] in na[0]:
                    head_index=head_list.index(item[1])
                    print(head_index)
                    print(item[3]+item[2])
                    na[head_index]=item[3]+item[2]
        for item in name_list:
            ws.append(item)
        wb.save(self.excel_name)

    def write_excel(self):
        head_list = ['姓名', '所在事务所', '出资日期/缴付期限', '出资额', '撤资日期', '变更日期']
        if not os.path.exists(self.excel_name):
            excel=Workbook()
            ws = excel.active
            ws.append(head_list)
        else:
            excel=load_workbook(self.excel_name)
            ws = excel.active
        time_list=[]
        name_dict={}
        item_list=[]
        name_list=[]
        print("len(name_list)={}".format(len(name_list)))
        for item in self.assets_list:
            name=item[0]
            time=item[1]
            if name not  in name_list:
                name_list.append(name)
            if name not  in name_dict.keys():
                name_dict[name]=[]
        print("print(len(name_list))={}".format(len(name_list)))

        for key in name_dict.keys():
            tmp_list = []
            for item in self.assets_list:
                name = item[0]
                time = item[1]
                mony = item[2]
                type = item[3]
                if name==key:
                    if [time,mony,type] not in tmp_list:
                        tmp_list.append([time,mony,type])
                    tmp_list.sort()
            name_dict[key]=tmp_list


        print(name_dict)
        print("len(name_dict)={}".format(len(name_dict)))

        for key,value in name_dict.items():
            final_list=['','','','','','']
            final_list[0] = key
            final_list[1] = self.asset_name
            final_list[2] = '/'.join(value[0][0].split('-'))
            if '/'.join(value[-1][0].split('-')) != self.time_list[-1]:
                final_list[4] = '/'.join(value[-1][0].split('-'))
            final_list[5]=';'.join(self.time_list)
            for va in value:
                if '入股' in va:
                    final_list[2]='/'.join(va[0].split('-'))
                    final_list[3]="货币"+va[1]+"万人民币"
                if '退出' in va:
                    final_list[4]='/'.join(va[0].split('-'))
            print(final_list)
            ws.append(final_list)
        excel.save(self.excel_name)
        shutil.copy(self.excel_name,self.final_name)


if __name__=="__main__":
    mt=MakeTransfer()
    mt.set_excel_name()
    #make_folder()
    mt.start_transfer()
    print("assets_list:{}".format(mt.assets_list))
    print("len(mt.assets_list):{}".format(len(mt.assets_list)))
    mt.write_excel()
    #mt.transfor_first()