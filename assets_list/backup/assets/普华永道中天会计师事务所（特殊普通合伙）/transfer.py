import json
import os
import re
import shutil
import sys

from openpyxl import Workbook, load_workbook


class MakeTransfer():
    def __init__(self):
        self.assets_list=[]
        self.type1_list=[]
        self.type2_list=[]
        self.type3_list = []
        self.time_list=[]
        self.excel_name=""
        self.change_list=[]

    def set_excel_name(self):
        dir_name=os.path.dirname(os.path.abspath(__file__))
        self.excel_name='E:\接单\code_3\股东变更.xlsx'
        self.asset_name=os.path.basename(dir_name)
        print(self.excel_name)

    def get_type_list(self):
        filist=os.listdir('.')
        for fi in filist:
            if fi.endswith('缴付期限.txt'):
                time=fi.split('_')[0]
                if time not in self.change_list:
                    self.change_list.append(time)
                self.type1_list.append(fi)
            elif fi.endswith('合伙人姓名或名称.txt'):
                time=fi.split('_')[0]
                if time not in self.change_list:
                    self.change_list.append(time)
                self.type2_list.append(fi)
            elif fi.endswith('出资方式.txt'):
                time=fi.split('_')[0]
                if time not in self.change_list:
                    self.change_list.append(time)
                self.type3_list.append(fi)
            else:
                print('pass file {}'.format(fi))
        print("type1_list:")
        print(self.type1_list)
        print("type2_list:")
        print(self.type2_list)
        print("type3_list:")
        print(self.type3_list)
        print("change_list:")
        print(self.change_list)

    def transfor_first(self,file1):
        print("----------------Start transfor_first-------------")
        berfor_list=[]
        after_list=[]
        inter_list=[]#交集
        union_list=[]#并集
        difference_a=[]#差集
        difference_b=[] # 差集
        with open(file1,'r',encoding='utf-8') as f:
            lines=f.readlines()
            n=len(lines)
            for item in lines[0].split(";"):
                p=re.search('(\D+)(\d+-\d+-\d+)\D+([\d | "."]+).*',item)
                if p:
                    name=p.group(1)
                    time=p.group(2)
                    amount=''.join(p.group(3).split(".")[:-1])
                    berfor_list.append((name,time,amount,'入股'))
                else:
                    print("Not find!")
            print(berfor_list)
            print("berfor length:",len(berfor_list))
            for item in lines[1].split(";"):
                p=re.search('(\D+)(\d+-\d+-\d+)\D+([\d | "."]+).*',item)
                if p:
                    name=p.group(1)
                    time=p.group(2)
                    amount=''.join(p.group(3).split(".")[:-1])
                    after_list.append((name,time,amount,'入股'))
                else:
                    print("Not find!")
            print(after_list)
            print("after_list length:", len(after_list))
            difference_a=set(berfor_list).difference(set(after_list))
            difference_b = set(after_list).difference(set(berfor_list))
            inter_list=set(berfor_list).intersection(set(after_list))
            union_list=set(berfor_list).union(set(after_list))
            # print(difference_a)
            # print(difference_b)
            print(inter_list)
            print("inter_list length:", len(inter_list))
            print(union_list)
            print("union_list length:", len(union_list))
            self.assets_list.extend(union_list)

    def transfor_second(self,file2,time):
        print("----------------Start transfor_second-------------")
        berfor_list=[]
        after_list=[]
        inter_list=[]#交集
        union_list=[]#并集
        difference_a=[]#差集
        difference_b=[] # 差集
        with open(file2,'r',encoding='utf-8') as f:
            lines=f.readlines()
            for na in lines:
                if '\n' in na:
                    na=na.replace('\n','')
                if '[新增]' in na:
                    name=na.replace('[新增]','')
                    time=time
                    amount=''
                    union_list.append((name,time,amount,'新增'))
                elif '[退出]' in na:
                    name=na.replace('[退出]','')
                    time=time
                    amount=''
                    union_list.append((name,time,amount,'退出'))
                else:
                    name=na
                    time=time
                    amount=''
                    union_list.append((name,time,amount,'持股'))
        print(union_list)
        self.assets_list.extend(union_list)

    def transfor_third(self,file3,time):
        print("----------------Start transfor_third-------------")
        berfor_list=[]
        after_list=[]
        inter_list=[]#交集
        union_list=[]#并集
        difference_a=[]#差集
        difference_b=[] # 差集
        line_list=[]
        new_lines=[]
        new_lines2=[]
        new_lines3 = []
        with open(file3,'r',encoding='utf-8') as f:
            lines=f.readlines()

        for index,item in enumerate(lines):
            # lines[index]=re.sub(';\s*[新增]','[新增]',item)
            # lines[index] = re.sub(';\s*[退出]', '[退出]', item)
            tmp_list=lines[index].split(';')
            new_lines.extend(tmp_list)
        for item in new_lines:
            if '\n' in item:
                new_lines2.append(item.replace('\n',''))
            else:
                new_lines2.append(item)
        for item in new_lines2:
            if item != ' ':
                new_lines3.append(item)
        print(new_lines3)
        print(len(new_lines3))
        for na in new_lines3:
            p=re.search('(\D+)(\d+).*',na)
            p2=re.search('(\D+).*[退出].*',na)
            p3 = re.search('(\D+).*[新增].*', na)
            if p or p2 or p3:
                if '[新增]' in na:
                    name = p3.group(1).replace('货币','') if '货币' in p3.group(1) else p3.group(1)
                    time = time
                    amount = 50
                    line_list.append((name, time, amount, '新增'))
                elif '[退出]' in na:
                    name =  p2.group(1).replace('货币','') if '货币' in p2.group(1) else p2.group(1)
                    time = time
                    amount =  50
                    line_list.append((name, time, amount, '退出'))
                else:
                    name = p.group(1).replace('货币','') if '货币' in p.group(1) else p.group(1)
                    time = time
                    amount = p.group(2)
                    line_list.append((name, time, amount, '持股'))
            else:
                print('In transfor_third,can not deal {}'.format(na))
        print("Transfor_third line list {}".format(line_list))
        self.assets_list.extend(line_list)

    def transfer_union(self):
        self.get_type_list()
        for fi1 in self.type1_list:
            self.transfor_first(fi1)

        for fi2 in self.type2_list:
            time=fi2.split('_')[0]
            self.time_list.append(time)
            # new_time=time[:4]+"-"+time[4:6]+"-"+time[6:]
            self.transfor_second(fi2,time)
        for fi3 in self.type3_list:
            time=fi3.split('_')[0]
            self.time_list.append(time)
            self.transfor_third(fi3,time)

    def write_excel2(self):
        wb=Workbook()
        ws=wb.active
        head_list=['姓名']
        tmp_list=[]
        name_list=[]
        for item in self.assets_list:
            if item[1] not  in tmp_list:
                tmp_list.append(item[1])
            if item[0] not in name_list:
                name_list.append([item[0]])
        sorted(tmp_list)
        head_list.extend(tmp_list)
        print(name_list)
        print("Totle {} people!".format(len(name_list)))
        ws.append(head_list)
        print(head_list)
        tmp_list2=['']*(len(head_list)-1)
        print(tmp_list2)
        for item2 in name_list:
            item2.extend(tmp_list2)
        print(len(head_list))
        print(len(name_list[0]))

        for item in self.assets_list:
            for na in name_list:
                print(na)
                print(item)
                if item[0] in na[0]:
                    head_index=head_list.index(item[1])
                    print(head_index)
                    print(item[3]+item[2])
                    na[head_index]=item[3]+item[2]
        for item in name_list:
            ws.append(item)
        wb.save(self.excel_name)

    def write_excel(self):
        head_list = ['姓名', '所在事务所', '出资日期/缴付期限', '出资额', '撤资日期', '变更日期']
        if not os.path.exists(self.excel_name):
            excel=Workbook()
            ws = excel.active
            ws.append(head_list)
        else:
            excel=load_workbook(self.excel_name)
            ws = excel.active
        time_list=[]
        name_dict={}
        item_list=[]
        name_list=[]
        print("len(name_list)={}".format(len(name_list)))
        for item in self.assets_list:
            name=item[0]
            time=item[1]
            if name not  in name_list:
                name_list.append(name)
            if name not  in name_dict.keys():
                name_dict[name]=[]
        print("print(len(name_list))={}".format(len(name_list)))

        for key in name_dict.keys():
            tmp_list = []
            for item in self.assets_list:
                name = item[0]
                time = item[1]
                mony = item[2]
                type = item[3]
                if name==key:
                    if [time,mony,type] not in tmp_list:
                        tmp_list.append([time,mony,type])
                    tmp_list.sort()
            name_dict[key]=tmp_list


        print(name_dict)
        print("len(name_dict)={}".format(len(name_dict)))

        for key,value in name_dict.items():
            final_list=['','','','','','']
            final_list[0] = key
            final_list[1] = self.asset_name
            final_list[5]=';'.join(self.change_list)
            for va in value:
                if '入股' in va:
                    final_list[2]='/'.join(va[0].split('-'))
                    final_list[3]="货币"+va[1]+"万人民币"
                if '退出' in va:
                    final_list[4]=va[0]
            print(final_list)
            ws.append(final_list)
        excel.save(self.excel_name)


if __name__=="__main__":
    mt=MakeTransfer()
    mt.set_excel_name()
    #make_folder()
    mt.transfer_union()
    print("assets_list:{}".format(mt.assets_list))
    print("len(mt.assets_list):{}".format(len(mt.assets_list)))
    mt.write_excel()
    #mt.transfor_first()